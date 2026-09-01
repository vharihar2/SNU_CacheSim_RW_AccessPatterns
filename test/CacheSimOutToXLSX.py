import sys
import os
import re
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def parse_and_export_cache_log(file_path, output_excel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cache State"

    # Define Styles
    font_bold_red = Font(name="Calibri", size=11, bold=True, color="FF0000") # D:T
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")     # MCA:T only
    font_red = Font(name="Calibri", size=11, color="FF0000")              # Default
    font_normal = Font(name="Calibri", size=11, color="555555")              # Default
    fill_invalid = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # V:F

    # Regex to match individual cache entry tokens
    entry_pattern = re.compile(r'(-?\d+,-?\d+)\s*\(V:([TF]),\s*D:([TF]),\s*MCA:([TF]),\s*TS:(-?\d+),\s*LAT:(-?\d+)\)')

    with open(file_path, 'r') as f:
        lines = f.readlines()

    ws.append(["Serial #", "L1 Entry 1", "L1 Entry 2", "L2 Entry 1", "L2 Entry 2", "L2 Entry 3", "L2 Entry 4"])

    for row_idx, line in enumerate(lines, start=2):
        if not line.strip():
            continue
        
        # Extract Serial Number
        serial_match = re.match(r'^\s*(\d+)\.', line)
        serial_num = serial_match.group(1) if serial_match else str(row_idx - 1)
        
        ws.cell(row=row_idx, column=1, value=int(serial_num)).alignment = Alignment(horizontal="center")
        
        # Find all cache entries in order
        entries = entry_pattern.findall(line)
        
        for col_idx, (addr_data, valid, dirty, mca, ts, lat) in enumerate(entries, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Formats address pair and all attributes including TS and LAT
            cell.value = f"{addr_data}\n(V:{valid}, D:{dirty}, MCA:{mca}, TS:{ts}, LAT:{lat})"
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
            
            # Formatting Rules
            if dirty == 'T':
                if mca == 'T':
                    cell.font = font_bold_red
                else:
                    cell.font = font_red
            elif dirty == 'F':
                if mca == 'T':
                    cell.font = font_bold
                else:
                    cell.font = font_normal
            else:
                cell.font = font_normal
                
            if valid == 'F':
                cell.fill = fill_invalid

    wb.save(output_excel)
    print(f"Successfully generated: {output_excel}")

def main():
    parser = argparse.ArgumentParser(description="Visualize Cache Log file into a styled Excel sheet.")
    parser.add_argument("input_file", help="Path to the input text log file")
    parser.add_argument("-o", "--output", help="Optional output XLSX file path", default=None)

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' does not exist.")
        sys.exit(1)

    output_file = args.output
    if not output_file:
        base_name = os.path.splitext(args.input_file)[0]
        output_file = f"{base_name}_visualized.xlsx"

    parse_and_export_cache_log(args.input_file, output_file)

if __name__ == "__main__":
    main()